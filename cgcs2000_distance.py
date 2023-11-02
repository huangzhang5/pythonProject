import openpyxl
import math
from tkinter.filedialog import askopenfilename, asksaveasfilename

# 打开Excel文件
input_file = askopenfilename(title="打开文件",
                             filetypes=[('Excel Files', '*.xlsx')],
                             initialfile="D:/研究生/各种任务/surfer")
if not input_file:
    print("未选择任何文件")
    exit()
workbook = openpyxl.load_workbook(input_file)

# 选择第一个工作表
sheet = workbook.active

# 遍历每行数据
for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
    # 获取点的坐标
    x = row[0]
    y = row[1]
    z = row[2]

    # 假定圆心坐标
    center_x = 530000
    center_y = 4370000
    center_z = -20

    # 计算距离
    distance = math.sqrt((x - center_x) ** 2 + (y - center_y) ** 2 + (z - center_z) ** 2)

    # 将距离值写入对应行的第四列
    sheet.cell(row=row_num, column=4).value = distance

# 保存修改后的Excel文件
output_file = asksaveasfilename(defaultextension='.xlsx',
                                filetypes=[("Excel Files", "*.xlsx")])
if not output_file:
    print("未选择保存文件位置")
    exit()
workbook.save(output_file)
